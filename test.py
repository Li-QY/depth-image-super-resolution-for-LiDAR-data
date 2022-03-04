import math
import os
import os.path as osp
import sys

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import torchvision
import yaml
import matplotlib
import matplotlib.cm as cm
import matplotlib.pyplot as plt
from math import ceil
from addict import Dict
from openpyxl import load_workbook
from torch.optim.lr_scheduler import MultiStepLR
from torch.utils.tensorboard import SummaryWriter
from tqdm import tqdm

import time
import torch.nn as nn
import torchvision.utils as utils
from openpyxl import load_workbook

from libs.modules.ssim import SSIM
from libs.modules.srgan import Generator, init_weights
from libs.datasets.mpo_crop import Augmentation, depth_to_ortho, revert_depth
from libs.modules.loss import ChamferLoss
from libs.modules.ssim import SSIM
from utils_aug import (
    get_device,
    test_dataloader,
    val_calcu,
    trans3D,
    _sample_topleft,
)

# generator, mode, init, utils, mpo
#

# category = {
#     "1": ["coast", 103],
#     "2": ["forest", 116],
#     "3": ["parkingin", 105],
#     "4": ["parkingout", 108],
#     "5": ["residential", 106],
#     "6": ["urban", 112],
# }


def scale(imgs):
    # [-1.0, 1.0] -> [0.0, 1.0]
    return (imgs + 1.0) / 2.0


def umiscale(imgs):
    return imgs * 2.0 - 1.0


def pad_image(image, mask, crop_size):
    new_h, new_w = image.shape[1:]
    pad_h = max(crop_size - new_h, 0)
    pad_w = max(crop_size - new_w, 0)
    padded_image = torch.FloatTensor(1, new_h + pad_h, new_w + pad_w).zero_()
    padded_mask = torch.FloatTensor(1, new_h + pad_h, new_w + pad_w).zero_()
    for i in range(1):  # RGB
        padded_image[:, [i], ...] = F.pad(
            image[:, [i], ...],
            pad=(0, pad_w, 0, pad_h),  # Pad right and bottom
            mode="constant",
            value=0,
        ).data
        padded_mask[:, [i], ...] = F.pad(
            mask[:, [i], ...],
            pad=(0, pad_w, 0, pad_h),  # Pad right and bottom
            mode="constant",
            value=0,
        ).data
    return padded_image, padded_mask


if __name__ == "__main__":
    # Load a yaml configuration file
    config = Dict(yaml.load(open(sys.argv[1]), Loader=yaml.SafeLoader))
    device = get_device(config.cuda)

    # Dataset
    test_loader = test_dataloader(config)
    augmentation = Augmentation(device, config.test)

    # Interpolation

    # Model setup
    n_ch = config.test.n_ch
    G = Generator(n_ch, n_ch, config.test.interp_factor)
    if config.test.gen_init is not None:
        print("Init:", config.test.gen_init)
        state_dict = torch.load(config.test.gen_init)
        G.load_state_dict(state_dict)
    # G.apply(init_weights)
    G.to(device)

    # Loss for test
    criterion_mse = nn.MSELoss().to(device)
    criterion_ssim = SSIM().to(device)
    criterion_cham = ChamferLoss().to(device)

    # precalcu
    vrange = val_calcu([384, 384])

    # Experiemtn ID
    exp_id = config.experiment_id

    crop_size = config.test.crop_size

    mse = 0.0
    ssim = 0.0
    psnr = 0.0
    cham = 0.0

    step = 0
    i = 1
    j = 1
    mode = exp_id

    # Tensorboard
    writer = SummaryWriter("runs/srgan_test/random_crop/" + exp_id)

    print("Experiment ID:", exp_id)

    for iteration, imgs_HR in tqdm(
        enumerate(test_loader, 1),
        desc="Test/Iteration",
        total=len(test_loader),
        leave=False,
    ):
        crop_start = _sample_topleft(config.test)
        sh, sw = crop_start
        # disp = imgs_HR["disp"]
        mask_HR = imgs_HR["mask"]
        imgs_HR_dep = imgs_HR["depth"]  # BxHxW
        # imgs_HR_ref = imgs_HR["refl"]
        pad_h, pad_w = imgs_HR_dep.shape[1:]
        stride_rate = 2 / 3.0
        stride = int(ceil(crop_size * stride_rate))
        h_grid = int(ceil((pad_h - crop_size) / float(stride)) + 1)
        w_grid = int(ceil((pad_w - crop_size) / float(stride)) + 1)
        count = 0
        for ih in range(h_grid):
            for iw in range(w_grid):
                sh, sw = ih * stride, iw * stride
                eh, ew = min(sh + crop_size, pad_h), min(sw + crop_size, pad_w)
                sh, sw = eh - crop_size, ew - crop_size  # Stay within image size
                image_sub = imgs_HR_dep[..., sh:eh, sw:ew]
                mask_sub = mask_HR[..., sh:eh, sw:ew].float()
                # image_sub, mask_sub = pad_image(image_sub, mask_sub, crop_size)

                image_sub_HR, image_sub_LR = augmentation(
                    image_sub, "depth", config.test.flip, crop_start
                )
                # imgs_HR_ref, imgs_LR_ref = augmentation(
                #     imgs_HR_ref, "refl", config.test.flip, crop_start
                # )
                mask_sub_HR, mask_sub_LR = augmentation(
                    mask_sub, "mask", config.test.flip, crop_start
                )
                # disp_HR, disp_LR = augmentation(disp, "disp", config.test.flip, crop_start)
                imgs_SR = G(scale(image_sub_LR))

                # imgs_SR = umiscale(revert_depth(scale(imgs_SR)))    , scale(imgs_LR_ref), mask_sub_LR

                # imgs_SR = F.interpolate(image_sub_LR, scale_factor=4, mode=mode)
                # batch_size = image_sub_HR.size(0)
                mse_SR = criterion_mse(image_sub_HR, imgs_SR)
                ssim_SR = criterion_ssim(image_sub_HR, imgs_SR).item()
                psnr_SR = 10 * torch.log10(1.0 / mse_SR).item()
                cham_SR = criterion_cham(image_sub_HR, imgs_SR, [sh, sw]).item()
                mse_SR = mse_SR.item()

                step = step + 1
                writer.add_scalar(mode + "/Scoremse", mse_SR, step)
                writer.add_scalar(mode + "/Scoressim", ssim_SR, step)
                writer.add_scalar(mode + "/Scorepsnr", psnr_SR, step)
                writer.add_scalar(mode + "/Scorecham", cham_SR, step)

                mse = mse_SR + mse
                ssim = ssim_SR + ssim
                psnr = psnr_SR + psnr
                cham = cham_SR + cham

                if step % 10 == 1:
                    imgs_LR_inter = F.interpolate(
                        image_sub_LR,
                        scale_factor=config.test.interp_factor,
                        mode="nearest",
                    )

                    ver_SR = trans3D(scale(imgs_SR) * 120, vrange)
                    ver_HR = trans3D(scale(image_sub_HR) * 120, vrange)
                    colors_tensor = torch.abs(scale(image_sub_HR) - scale(imgs_SR))
                    colors_tensor = colors_tensor.view(1, 1, 384, 384)
                    colors_tensor = colors_tensor.detach().cpu().numpy()

                    for array in colors_tensor:
                        array = array.squeeze()  # remove the redundant dim
                        colored_array = cm.jet(array)  # returns as RGBA
                        colored_array = colored_array[..., :3]  # only uses RGB

                    colors_tensor = colored_array.reshape(1, -1, 3) * 255.0
                    colors_tensor = torch.from_numpy(colors_tensor)
                    writer.add_mesh(
                        mode, vertices=ver_SR, colors=colors_tensor, global_step=step,
                    )
                    writer.add_mesh(
                        "HR", vertices=ver_HR, colors=colors_tensor, global_step=step,
                    )

                    colored_array = colored_array.transpose(2, 0, 1)
                    summary = [imgs_LR_inter, image_sub_HR, imgs_SR]
                    summary = torch.cat(summary, dim=3)
                    summary = torch.clamp(scale(summary), 0.0, 1.0)
                    writer.add_images(mode + "Results", summary, step)
                    writer.add_image(mode + "colormapping", colored_array, step)

    wb = load_workbook("sta/test_score.xlsx")
    ws = wb["Sheet1"]
    ws.append(
        [
            mode,
            float(mse / 950.0),
            float(ssim / 950.0),
            float(psnr / 950.0),
            float(cham / 950.0),
        ]
    )
    wb.save("sta/test_score.xlsx")
    print("All is done")
