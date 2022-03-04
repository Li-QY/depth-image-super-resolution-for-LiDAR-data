import math
import os
import os.path as osp
import sys

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import torchvision
import yaml
import matplotlib
import matplotlib.cm as cm
import matplotlib.pyplot as plt
from addict import Dict
from openpyxl import load_workbook
from torch.optim.lr_scheduler import MultiStepLR
from torch.utils.tensorboard import SummaryWriter
from tqdm import tqdm

import time
import torch.nn as nn
import torchvision.utils as utils
from openpyxl import load_workbook

from libs.modules.ssim import SSIM
from libs.modules.DF_mask import Generator, init_weights
from pointnet.pointnet import PointNetSeg
from libs.datasets.mpo_crop import Augmentation, midPointSet
from libs.modules.loss import ChamferLoss
from libs.modules.ssim import SSIM
from utils_aug import (
    get_device,
    test_dataloader,
    val_calcu,
    trans3D,
    _sample_topleft,
    makeTestBlock,
)

# generator, mode, init, utils, mpo

category = {
    "1": ["coast", 103],
    "2": ["forest", 116],
    "3": ["parkingin", 105],
    "4": ["parkingout", 108],
    "5": ["residential", 106],
    "6": ["urban", 112],
}


def scale(imgs):
    # [-1.0, 1.0] -> [0.0, 1.0]
    return (imgs + 1.0) / 2.0


def umiscale(imgs):
    return imgs * 2.0 - 1.0


if __name__ == "__main__":
    # Load a yaml configuration file
    config = Dict(yaml.load(open(sys.argv[1]), Loader=yaml.SafeLoader))
    device = get_device(config.cuda)

    # Dataset
    test_loader = test_dataloader(config)
    augmentation = Augmentation(device, config.test)

    # Interpolation

    # Model setup
    n_ch = config.test.n_ch
    G = Generator(n_ch, n_ch, config.test.interp_factor)
    if config.test.gen_init is not None:
        print("Init:", config.test.gen_init)
        state_dict = torch.load(config.test.gen_init)
        G.load_state_dict(state_dict)
    # G.apply(init_weights)
    G.to(device)
    PG = PointNetSeg()
    if config.pointnet.gen_init is not None:
        print("Init:", config.pointnet.gen_init)
        state_dict = torch.load(config.pointnet.gen_init)
        PG.load_state_dict(state_dict)
    PG.to(device)

    # Loss for test
    criterion_mse = nn.MSELoss().to(device)
    criterion_ssim = SSIM().to(device)
    criterion_cham = ChamferLoss().to(device)

    # precalcu
    vrange = val_calcu([384, 384])

    # Experiemtn ID
    exp_id = config.experiment_id

    pmax = 0
    pmin = 147456
    ptotal = []

    mse = 0.0
    ssim = 0.0
    psnr = 0.0
    cham = 0.0

    mse_total = 0.0
    ssim_total = 0.0
    psnr_total = 0.0
    cham_total = 0.0

    step = 0
    i = 1
    j = 1
    mode = "DF_cham+mask"
    categoryNow = "coast"

    # Tensorboard
    writer = SummaryWriter("runs/srgan_test/" + exp_id)

    print("Experiment ID:", exp_id)

    for iteration, imgs_HR in tqdm(
        enumerate(test_loader, 1),
        desc="Test/Iteration",
        total=len(test_loader),
        leave=False,
    ):
        crop_start = _sample_topleft(config.test)
        sh, sw = crop_start
        mask_HR = imgs_HR["mask"]
        imgs_HR_dep = imgs_HR["depth"]
        imgs_HR_ref = imgs_HR["refl"]
        imgs_HR_dep, imgs_LR_dep = augmentation(
            imgs_HR_dep, "depth", config.test.flip, crop_start
        )
        imgs_HR_ref, imgs_LR_ref = augmentation(
            imgs_HR_ref, "refl", config.test.flip, crop_start
        )
        mask_HR, mask_LR = augmentation(mask_HR, "mask", config.test.flip, crop_start)

        # Generate fake images
        imgs_SR = G(scale(imgs_LR_dep), scale(imgs_LR_ref), mask_LR)
        mse_SR = criterion_mse(imgs_HR_dep, imgs_SR)
        ssim_SR = criterion_ssim(imgs_HR_dep, imgs_SR)
        psnr_SR = 10 * torch.log10(1.0 / mse_SR)
        # PointNet
        SRlabels, SRblocks, targets = makeTestBlock(imgs_SR, imgs_HR_dep, crop_start)
        pointset = midPointSet(SRlabels, SRblocks, targets)
        point_loader = torch.utils.data.DataLoader(  # Batch_of_PNx3x4096
            pointset,
            batch_size=config.pointnet.batch_size,
            shuffle=True,
            # num_workers=0,
            drop_last=True,
        )
        SRrecovery = []
        HRrecovery = []

        for points in point_loader:
            SRblocks = points["SR"]
            HRblocks = points["HR"]
            labels = points["label"]
            PointsSR = PG(SRblocks)
            SRrecovery.append(PointsSR.detach())
            HRrecovery.append(HRblocks.detach())

        # imgs_SR = F.interpolate(imgs_LR, scale_factor=interp_factor, mode=mode)

        # cham_SR = [
        #     criterion_cham(HRrecovery[i], SRrecovery[i], [sh, sw])
        #     for i in range(len(SRrecovery))
        # ]
        cham_SR = criterion_cham(torch.cat(HRrecovery), torch.cat(SRrecovery), [sh, sw])

        SRrecovery = torch.cat(SRrecovery).view(1, -1, 3)
        HRrecovery = torch.cat(HRrecovery).view(1, -1, 3)

        step = step + 1
        writer.add_scalar(mode + "/Scoremse", mse_SR, step)
        writer.add_scalar(mode + "/Scoressim", ssim_SR, step)
        writer.add_scalar(mode + "/Scorepsnr", psnr_SR, step)
        writer.add_scalar(mode + "/Scorecham", cham_SR, step)

        mse = mse_SR.detach().cpu() + mse
        ssim = ssim_SR.detach().cpu() + ssim
        psnr = psnr_SR.detach().cpu() + psnr
        cham = cham_SR.detach().cpu() + cham

        point_number = torch.sum((imgs_HR_dep + 1.0) / 2.0 != 0.0000)
        ptotal.append(point_number.cpu().numpy())
        if point_number > pmax:
            pmax = point_number
        elif point_number < pmin:
            pmin = point_number

        if i == 1:
            imgs_LR_inter = F.interpolate(
                imgs_LR_dep, scale_factor=config.test.interp_factor, mode="nearest"
            )

            # ver_SR = trans3D(scale(imgs_SR) * 120, vrange)
            # ver_HR = trans3D(scale(imgs_HR_dep) * 120, vrange)
            colors_tensor = torch.abs(HRrecovery - SRrecovery)
            # colors_tensor = colors_tensor.view(1, 1, 384, 384)
            colors_tensor = colors_tensor.detach().cpu().numpy()

            for array in colors_tensor:
                array = array.squeeze()  # remove the redundant dim
                colored_array = cm.jet(array)  # returns as RGBA
                colored_array = colored_array[..., :3]  # only uses RGB

            colors_tensor = colored_array.reshape(1, -1, 3) * 255.0
            colors_tensor = torch.from_numpy(colors_tensor)
            writer.add_mesh(
                mode, vertices=SRrecovery, colors=colors_tensor, global_step=step,
            )
            writer.add_mesh(
                "HR", vertices=HRrecovery, colors=colors_tensor, global_step=step,
            )

            # colored_array = colored_array.transpose(2, 0, 1)
            summary = [imgs_LR_inter, imgs_HR_dep, imgs_SR]
            summary = torch.cat(summary, dim=2)
            summary = torch.clamp(scale(summary), 0.0, 1.0)
            writer.add_images(mode + "Results", summary, step)
            # writer.add_image(mode + "colormapping", colored_array, step)
        i = i + 1
        if (i > category[str(j)][1]) & (j != 7):
            mse_total = mse_total + mse.cpu().numpy()
            ssim_total = ssim_total + ssim.cpu().numpy()
            psnr_total = psnr_total + psnr.cpu().numpy()
            cham_total = cham_total + cham.cpu().numpy()
            wb = load_workbook("sta/score.xlsx")
            ws = wb["Sheet2"]
            ws.append(
                [
                    mode,
                    categoryNow,
                    str(mse.cpu().numpy() / np.float(category[str(j)][1])),
                    str(ssim.cpu().numpy() / np.float(category[str(j)][1])),
                    str(psnr.cpu().numpy() / np.float(category[str(j)][1])),
                    str(cham.cpu().numpy() / np.float(category[str(j)][1])),
                ]
            )
            wb.save("sta/score.xlsx")
            print(categoryNow + " is finished")
            psnr = 0.0
            mse = 0.0
            ssim = 0.0
            cham = 0.0
            i = 1
            j = j + 1
            if j != 7:
                categoryNow = category[str(j)][0]
    wb = load_workbook("sta/score.xlsx")
    ws = wb["Sheet1"]
    ws.append(
        [
            mode,
            categoryNow,
            str(mse_total / 650.0),
            str(ssim_total / 650.0),
            str(psnr_total / 650.0),
            str(cham_total / 650.0),
        ]
    )
    wb.save("sta/score.xlsx")
    print("All is done")

    print("pmin is", pmin)
    print("pmax is", pmax)
    print("average is", np.average(ptotal))
    np.save("HR_number.npy", ptotal)
